# -*- coding: utf-8 -*-
"""
Created on Wed Jan 12 10:58:28 2022

@author: Hilal YILMAZ
"""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from collections import namedtuple
from docplex.mp.model import Model

def cspm(loc,obj,time, **kwargs):
    
    # Open Workbook and worksheet
    wb = load_workbook(loc, keep_vba=True, data_only=True)
    ws = wb.active


    # Access Data
    n =  ws["B7"].value          # of charging stations
    Bmin=ws["B5"].value/100      # Min SOC level (%)
    l =  ws["B6"].value          # route length (kWh): same as the energy required to reach the destination (e_n+1)
    T =  ws["B8"].value          # Non-stop travel time: : same as the time required to reach the destination (s_n+1)
    B =  ws["B9"].value          # energy capacity of EV (kWh)
    r0 = ws["B10"].value         # remaining energy at origin (kWh)
    bp = ws["B14"].value/100     # Nonlinear charging time/brakepoints (Energy-time)
    sp = ws["B15"].value/100     # Nonlinear charging time/slopes (Energy-time)

    CS = []  # Charging station data
    for i in range(7, 7+n):
        cs = []
        for j in range(4, 9):
            col = get_column_letter(j)
            cs.append(ws[col+str(i)].value)
        cs = tuple(cs)
        CS.append(cs)
    
    #Insert origin and destination nodes as dummy CSs
    CS.insert(0,(0,0,0,0,0))     #origin 
    CS.insert(n+1,(n+1,l,T,0,0)) #destination
           
    
    # Sets
    CS_info = namedtuple("CS", ["order", "energy", "time","power","cost"])    
    cs = [CS_info(*i) for i in CS]
    cs = {i: cs[i] for i in range(0, n+2)}

    # print("\n**CS data information**\n")    
    # for i in range(len(cs)):
    #     print(cs[i])  
    # print("\n**The first and last CSs are the OD nodes included as dummy CSs**\n\n")

    # Model
    mdl = Model(name="cspm", **kwargs)
    if time: mdl.set_time_limit(time)

    # Binary Decision Variables
    x = {i: mdl.binary_var(name="x%d" % (i)) for i in range(0, n+2)}
    z = {(i, j): mdl.binary_var(name="z%d.%d" % (i, j)) for i in range(1, n+2) for j in range(i)}
    b = {(i, j): mdl.binary_var(name="b%d.%d" % (i, j)) for i in range(1, n+2) for j in range(i)}
      
    # Continuous Decision Variables
    Ra = {i: mdl.continuous_var(ub=B, lb=B*Bmin, name="Ra%d" % (i)) for i in range(1, n+2)}
    Rd = {i: mdl.continuous_var(ub=B, lb=B*Bmin, name="Rd%d" % (i)) for i in range(0, n+1)}
    Ta = {i: mdl.continuous_var(lb=0, name="Ta%d" % (i)) for i in range(1, n+2)}
    Td = {i: mdl.continuous_var(lb=0, name="Td%d" % (i)) for i in range(0, n+1)}
     
    # Objective Function
    if obj== "traveltime": 
        objective = Ta[n+1]
    elif obj== "cost":
        objective = mdl.sum(cs[i].cost*(Td[i]-Ta[i]) for i in range(1, n+1))

    elif obj=="number_of_stops":
        objective = mdl.sum(x[i] for i in range(1, n+1))
                    
    if obj=="multi":
        mdl.set_multi_objective("min",[mdl.sum(cs[i].cost*(Td[i]-Ta[i]) for i in range(1, n+1)), Ta[n+1]])    
    else:
        mdl.minimize(objective)

    
    # C1 (Eq. 4): EV must be charged at a CS that can be reached with the initial remaining energy (r0)
    subset = [cs[i].order for i in cs if cs[i].energy <= r0-B*Bmin and i>0] 
    c1 = mdl.add(mdl.sum(x[i] for i in subset) >= 1)
    #print(c1)

    # C2 (Eq. 5): EV must be charged at a CS that requires energy less than the energy capacity (B) to reach the destination
    subset2 = [cs[i].order for i in cs if cs[n+1].energy-cs[i].energy <= B*(1-Bmin) and i<=n]
    c2 = mdl.add(mdl.sum(x[i] for i in subset2) >= 1)

    # C3 (Eqs. 6-7): If the EV is charged both at i and j, z equals to 1
    for i in range(1, n+2):
        for j in range(i):
            c3  = mdl.add(x[i]+x[j]-1<=z[i,j])
            c3_ = mdl.add(x[i]+x[j]>=2*z[i,j])


    # C4 (Eqs. 8-9): If the EV is charged at CS between z[i,j]=1 and z[j,k] =1, then b=1
    for i in range(1, n+2):
        for k in range(i):
            c4  = mdl.add(mdl.sum(z[i,j] for j in range(k+1,i))<=n*b[i,k])
            c4_ = mdl.add(mdl.sum(z[i,j] for j in range(k+1,i))>=b[i,k])

    #C5 (Eq. 10-11): The energy required to travel between two consecutive CSs  must not exceed the energy of the EV      
    for i in range(1, n+2):
        for j in range(i):           
            c5 = mdl.add(Rd[j] - Ra[i] - cs[i].energy + cs[j].energy <= l * (1-z[i,j]+b[i,j]))
            c5_= mdl.add(Rd[j] - Ra[i] - cs[i].energy + cs[j].energy >= -l * (1 - z[i,j]+b[i,j]))
 
    #C6 (Eq. 12): Time Balance                  
    for i in range(1, n+2):
        C7=mdl.add(cs[i].time-cs[i-1].time==Ta[i]-Td[i-1])
 
    #C8 (Eq. 13): Piecewise Linear Function: f(desired energy)=time (min) needed for charging to reach desired energy
    for i in range(1,n+1):
        ctime=mdl.piecewise(0, [(0,0),(B*bp,60*B*bp/cs[i].power)], 60/(cs[i].power*sp)) #the piecewise function     
        k8=mdl.add(Td[i]-Ta[i]==ctime(Rd[i])-ctime(Ra[i])) 
        
    #C9 (Eq. 14) is included as a lower bound when defining the variables, if lower bound is set to default (zero), activate C9
    # for i in range(1,n+1): 
    #     mdl.add(Ra[i]>=B*Bmin*x[i]) 
     
    #K10 (Eq. 15-16) If the EV is not charged at the CS, then x[i]=0
    for i in range(1,n+1):
        k10 = mdl.add(Rd[i]-Ra[i]<=B*x[i])
        k11 = mdl.add(Ra[i]-Rd[i]<=B*x[i])
    
    #K12 (Eq. 17-18)  If the EV does not spend time at the CS, then x[i]=0
    for i in range(1,n+1):
        k12 = mdl.add(Td[i]-Ta[i]<=60*(B*(1-bp)/cs[i].power+B*bp/(sp*cs[i].power))*x[i])  #ctime(B)
        k13 = mdl.add(Ta[i]-Td[i]<=60*(B*(1-bp)/cs[i].power+B*bp/(sp*cs[i].power))*x[i])
    
    #K14 (Eq. 19): The EV must visit (depart from) the origin
    mdl.add(x[0]==1)
    
    #K15 (Eq. 20): The EV must visit (arrive at) the destination
    mdl.add(x[n+1]==1)
    
    #K16 (Eq. 21): The EV must depart from the origin with the initial energy amount (r0)
    mdl.add(Rd[0]==r0)
    
    #K17 (Eq. 22): The departure time is zero at the beginning of the travel
    mdl.add(Td[0]==0)

  
    #Define KPI's
    tottime=Ta[n+1]
    mdl.add_kpi(tottime, publish_name="Travel time")
    
    totcost=mdl.sum(cs[i].cost*(Td[i]-Ta[i]) for i in range(1, n+1))
    mdl.add_kpi(totcost, publish_name="Cost")
    
    stops=mdl.sum(x[i] for i in range(1, n+1))
    mdl.add_kpi(stops, publish_name="Stops")
       
    return mdl
